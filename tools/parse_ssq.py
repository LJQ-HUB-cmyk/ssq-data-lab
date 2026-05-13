#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
将双色球历史开奖数据（Tab 分隔 txt 或 xlsx）解析为前端可加载的 JSON。

用法:
  python tools/parse_ssq.py 输入.txt [输出.json]
  python tools/parse_ssq.py 输入.xlsx [输出.json]

依赖:
  - txt: 仅标准库
  - xlsx: openpyxl (pip install openpyxl)

同时写出 data/draws.js (window.__SSQ_DATA__ = ...) 便于 file:// 直开。
"""

from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

_DATA_LINE_RE = re.compile(r"^\d+\t\d{7}\t\d{4}\t")


def _to_int(s):
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _to_date(v):
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return s


def _build_draw(issue, year, date, reds, blue):
    if not issue or any(r is None for r in reds) or blue is None:
        return None
    if len(reds) != 6:
        return None
    if not (1 <= blue <= 16):
        return None
    if any(r < 1 or r > 33 for r in reds):
        return None
    if len(set(reds)) != 6:
        return None
    return {
        "issue": str(issue),
        "year": year,
        "date": date,
        "reds": sorted(reds),
        "blue": blue,
    }


def parse_txt(path: str) -> list[dict]:
    draws = []
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.rstrip("\n").rstrip("\r")
            if not _DATA_LINE_RE.match(line):
                continue
            parts = line.split("\t")
            if len(parts) < 11:
                continue
            draw = _build_draw(
                issue=parts[1].strip(),
                year=_to_int(parts[2]),
                date=_to_date(parts[3]),
                reds=[_to_int(x) for x in parts[4:10]],
                blue=_to_int(parts[10]),
            )
            if draw:
                draws.append(draw)
    return draws


def parse_xlsx(path: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit(
            "解析 xlsx 需要 openpyxl，请先安装：pip install openpyxl"
        ) from e

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb["开奖数据"] if "开奖数据" in wb.sheetnames else wb.active

    rows: Iterable[tuple] = sheet.iter_rows(values_only=True)
    header = None
    for row in rows:
        if row and any(c == "期号" for c in row):
            header = list(row)
            break
    if not header:
        raise ValueError("xlsx 未找到包含『期号』的表头行")

    required = [
        "期号",
        "年份",
        "开奖日期",
        "红球1",
        "红球2",
        "红球3",
        "红球4",
        "红球5",
        "红球6",
        "蓝球",
    ]
    missing = [c for c in required if c not in header]
    if missing:
        raise ValueError(f"xlsx 缺少列：{missing}")

    idx = {name: header.index(name) for name in required}
    draws = []
    for row in rows:
        if not row or row[idx["期号"]] is None:
            continue
        reds = [_to_int(row[idx[f"红球{i}"]]) for i in range(1, 7)]
        draw = _build_draw(
            issue=str(row[idx["期号"]]).strip(),
            year=_to_int(row[idx["年份"]]),
            date=_to_date(row[idx["开奖日期"]]),
            reds=reds,
            blue=_to_int(row[idx["蓝球"]]),
        )
        if draw:
            draws.append(draw)
    return draws


def _write_outputs(output_json: Path, doc: dict) -> None:
    output_json.parent.mkdir(parents=True, exist_ok=True)
    compact = json.dumps(doc, ensure_ascii=False, separators=(",", ":"))
    output_json.write_text(compact, encoding="utf-8")

    draws_js = output_json.with_name("draws.js")
    draws_js.write_text("window.__SSQ_DATA__=" + compact, encoding="utf-8")


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(
            "用法：python tools/parse_ssq.py 输入(.txt|.xlsx) [输出.json]",
            file=sys.stderr,
        )
        return 2

    input_path = argv[1]
    output_path = Path(argv[2] if len(argv) >= 3 else "data/draws.json")

    if not os.path.exists(input_path):
        raise FileNotFoundError(input_path)

    ext = os.path.splitext(input_path)[1].lower()
    draws = (
        parse_xlsx(input_path) if ext in {".xlsx", ".xlsm"} else parse_txt(input_path)
    )

    doc = {
        "meta": {
            "source": os.path.basename(input_path),
            "count": len(draws),
            "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        },
        "draws": draws,
    }
    _write_outputs(output_path, doc)
    print(f"OK parsed {len(draws)} draws -> {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
