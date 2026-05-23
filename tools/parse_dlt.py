#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
解析大乐透历史 txt / xlsx 数据，输出标准 draws.json 结构。

数据格式 (输入文件每行)：
  期号 [日期] f1 f2 f3 f4 f5 b1 b2
  示例: 25001 2025-01-04 03 11 15 22 30 04 09
       (日期可选，date 可空)

输出 (data/dlt-draws.json)：
  {
    "meta": { "source": "...", "count": N, "generatedAt": "..." },
    "draws": [
      { "issue": "25001", "year": 2025, "date": "2025-01-04",
        "front": [3, 11, 15, 22, 30], "back": [4, 9] }, ...
    ]
  }

用法:
  python tools/parse_dlt.py 历史.txt data/dlt-draws.json
  python tools/parse_dlt.py 历史.xlsx data/dlt-draws.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

# 大乐透：前区 5 选 35，后区 2 选 12
FRONT_SIZE = 35
FRONT_PICK = 5
BACK_SIZE = 12
BACK_PICK = 2

# 期号 5 位（如 26054）
_ISSUE_RE = re.compile(r"^\s*(\d{5})\b")
# 形如 "2026-05-21" 的日期
_DATE_RE = re.compile(r"\b(20\d{2})-(\d{2})-(\d{2})\b")


def issue_year(issue: str) -> int:
    """从 5 位期号推年份 (前 2 位为年份后两位)。"""
    yy = int(issue[:2])
    # 大乐透 2007 年才开售，所以 yy < 7 视为 21 世纪 (但保险起见全部按 2000+)。
    # 我们项目数据从 2007 起，所有 yy 都是 07 - 99。
    return 2000 + yy if yy < 80 else 1900 + yy


def validate_draw(d: dict) -> None:
    if not re.fullmatch(r"\d{5}", d["issue"]):
        raise ValueError(f"bad issue: {d['issue']}")
    front = d["front"]
    if len(front) != FRONT_PICK or len(set(front)) != FRONT_PICK:
        raise ValueError(f"bad front: {front}")
    if any(n < 1 or n > FRONT_SIZE for n in front):
        raise ValueError(f"front out of range: {front}")
    back = d["back"]
    if len(back) != BACK_PICK or len(set(back)) != BACK_PICK:
        raise ValueError(f"bad back: {back}")
    if any(n < 1 or n > BACK_SIZE for n in back):
        raise ValueError(f"back out of range: {back}")


def parse_text_lines(lines):
    """逐行解析 'issue [date] f1..f5 b1 b2'。"""
    draws = []
    for raw in lines:
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        m = _ISSUE_RE.match(line)
        if not m:
            continue
        issue = m.group(1)
        date_match = _DATE_RE.search(line)
        date = date_match.group(0) if date_match else None
        # 抽取所有 1-2 位整数（去掉日期里的数字会和号码混淆，先把日期替换掉再扫）
        rest = line.replace(issue, "", 1)
        if date_match:
            rest = rest.replace(date_match.group(0), "", 1)
        nums = [int(x) for x in re.findall(r"\b\d{1,2}\b", rest)]
        if len(nums) < FRONT_PICK + BACK_PICK:
            continue
        front = sorted(nums[:FRONT_PICK])
        back = sorted(nums[FRONT_PICK : FRONT_PICK + BACK_PICK])
        draw = {
            "issue": issue,
            "year": issue_year(issue),
            "date": date,
            "front": front,
            "back": back,
        }
        try:
            validate_draw(draw)
        except ValueError:
            continue
        draws.append(draw)
    return draws


def parse_text_file(path: Path):
    with path.open("r", encoding="utf-8", errors="replace") as fp:
        return parse_text_lines(fp)


def parse_xlsx_file(path: Path):
    """xlsx 解析使用 openpyxl，若不可用则降级为 zip+xml 解析。"""
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return _parse_xlsx_naive(path)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_text = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        rows_text.append(" ".join(cells))
    return parse_text_lines(rows_text)


def _parse_xlsx_naive(path: Path):
    """openpyxl 不可用时的兜底：直接读 sharedStrings + sheet1。仅做最尽力解析。"""
    import zipfile
    from xml.etree import ElementTree as ET

    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(path) as zf:
        strings = []
        if "xl/sharedStrings.xml" in zf.namelist():
            xml = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in xml.findall("s:si", ns):
                texts = [t.text or "" for t in si.iter() if t.tag.endswith("}t")]
                strings.append("".join(texts))
        sheet_name = next(
            (n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")),
            None,
        )
        if not sheet_name:
            return []
        sheet_xml = ET.fromstring(zf.read(sheet_name))
        rows_text = []
        for row in sheet_xml.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row"):
            cells = []
            for c in row.findall("s:c", ns):
                t = c.attrib.get("t")
                v = c.find("s:v", ns)
                if v is None:
                    cells.append("")
                    continue
                if t == "s":
                    idx = int(v.text or 0)
                    cells.append(strings[idx] if idx < len(strings) else "")
                else:
                    cells.append(v.text or "")
            rows_text.append(" ".join(cells))
        return parse_text_lines(rows_text)


def write_outputs(target: Path, draws: list, source_label: str):
    target.parent.mkdir(parents=True, exist_ok=True)
    doc = {
        "meta": {
            "source": source_label,
            "count": len(draws),
            "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        },
        "draws": sorted(draws, key=lambda x: x["issue"]),
    }
    compact = json.dumps(doc, ensure_ascii=False, separators=(",", ":"))
    target.write_text(compact, encoding="utf-8")
    target.with_name("dlt-draws.js").write_text(
        "window.__DLT_DATA__=" + compact, encoding="utf-8"
    )


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("input", type=Path, help="输入 txt 或 xlsx")
    p.add_argument("output", type=Path, help="输出 dlt-draws.json")
    p.add_argument("--source-label", default="manual", help="meta.source 字段")
    args = p.parse_args(argv)

    if not args.input.exists():
        print(f"input not found: {args.input}", file=sys.stderr)
        return 2
    suffix = args.input.suffix.lower()
    if suffix == ".xlsx":
        draws = parse_xlsx_file(args.input)
    else:
        draws = parse_text_file(args.input)
    if not draws:
        print("parse_empty: no draws parsed", file=sys.stderr)
        return 3
    write_outputs(args.output, draws, args.source_label)
    print(f"parsed {len(draws)} draws -> {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
